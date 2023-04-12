import time as t
import os
import re
import sys
import string
import openpyxl as xl
from openpyxl.styles import Font, fonts
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from variables import *


# VARIABLES #


current_date = t.localtime()
current_year = current_date[0]
current_month = current_date[1]
homedir = os.getcwd()
systempath = homedir

if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    systempath = sys._MEIPASS


# COMMON FUNCTIONS


def ts():
    # Time (and date) stamp. Used in log function

    dtstp = t.localtime()
    c_date = f"{dtstp[2]}/{dtstp[1]}/{dtstp[0]}"
    c_time = f"{dtstp[3]:0>2}:{dtstp[4]:0>2}:{dtstp[5]:0>2}"
    return [c_time, c_date]


def log(message, nots=False):
    # Simple log function

    dtstp = current_date
    c_date = f"{dtstp[2]:0>2}_{dtstp[1]:0>2}_{dtstp[0]:0>4}"
    c_time = f"{dtstp[3]:0>2}h{dtstp[4]:0>2}m"

    try:
        os.chdir("log")
    except FileNotFoundError:
        os.mkdir("log")
        os.chdir("log")
    filename = f"{c_date}_{c_time}.log"

    text = []
    try:
        with open(filename, mode="r", encoding="utf-8") as fout:
            for _ in fout.readlines():
                text_line = re.sub(r"\n|\d+:\d+:\d+:\s", "", _)
                text.append(text_line)
    except FileNotFoundError:
        pass
    finally:
        text.append(message)
    with open(filename, mode="w", encoding="utf-8") as fin:
        for line in text:
            if nots:
                fin.write(f"{line}\n")
            else:
                fin.write(f"{ts()[0]}: {line}\n")
    os.chdir("..")


def benchmark(func):
    # Wrapper for benchmark test

    def wrap(*args, **kwargs):
        start = t.time()
        result = func(*args, **kwargs)
        end = t.time()
        log(f"[BENCHMARK] of [{func.__name__}]: {end - start:.3f} seconds")
        return result
    return wrap


def months (month_id):
    return month_text_dict[month_id - 1]


def convert_date(input_datetime):
    result = False
    year = input_datetime.year
    month = input_datetime.month
    day = input_datetime.day
    result = [year, month, day]
    return result


def save_to_file(array=[], filename="history.txt"):
    # Save [] as (strings + \n) in filename

    os.chdir(systempath)
    with open(filename, mode="w", encoding="utf-8") as file:
        for line in array:
            file.write(f"{line}\n")
    os.chdir(homedir)


def read_from_file(filename="history.txt"):
    # Load strings from filename, \n excluded

    os.chdir(systempath)
    output = []
    with open(filename, mode="r", encoding="utf-8") as file:
        for line in file.readlines():
            string = re.sub(r"\n", r"", line)
            output.append(eval(string))
    os.chdir(homedir)
    return output


def get_files(directory=os.getcwd()):
    # Get list of files in directory

    os.chdir(systempath)
    files = os.listdir(path=directory)
    os.chdir(homedir)
    return files


def parse_version():
    result = False
    version = str(len(get_files('log')))
    version_d = re.split(r"\B", version)
    version_w = string.ascii_lowercase[int(version_d[-1])]
    del version_d[-1]
    result = "0." + ".".join(version_d) + version_w
    return result

# BOT FUNCTIONS #


def found_spaces(input_string):
    result = False
    if input_string:
        result = re.findall(r"\s", input_string)
    return result


def sub_marked(input_string):
    found = re.findall(r"\u2705", input_string)
    if found:
        input_string = re.sub(r"\u2705", r"Доход ", input_string)
    return input_string


def sub_spaces(input_string):
    found = re.findall(r"\s\s", input_string)
    while len(found) > 0:
        input_string = re.sub(r"\s\s", r" ", input_string)
        found = re.findall(r"\s\s", input_string)
    return input_string


def sub_nlines(input_string):
    found = re.findall(r"\n", input_string)
    while len(found) > 0:
        input_string = re.sub(r"\n", r" ", input_string)
        found = re.findall(r"\n", input_string)
    return input_string


def sub_points(input_string):
    input_string = re.sub(r",", r".", input_string)
    return input_string


def check_values(input_string=""):
    result = [False, False]
    input_string = sub_marked(input_string)
    if found_spaces(input_string):
        input_string = sub_nlines(input_string)
        input_string = sub_spaces(input_string)
        input_string = sub_points(input_string)
        input_string = sub_marked(input_string)
        words_and_digits = re.findall(r"\b\w+[,.]\w+\b|\b\w+\b", input_string)
        words = []
        for _ in words_and_digits:
            try:
                value = float(_)
                result[1] = value
            except ValueError as e:
                words.append(_)
        result[0] = " ".join(words)
        if not re.findall(r"доход", result[0], flags=re.IGNORECASE):
            result[1] = 0 - result[1]
    return result


def clear_file(filename="history.txt"):
    os.chdir(systempath)
    with open(filename, mode="w", encoding="utf-8") as file:
        file.write(f"")
    os.chdir(homedir)


def clear_last(filename="history.txt"):
    os.chdir(systempath)
    output = read_from_file(filename)
    output = output[:-1]
    save_to_file(output)
    os.chdir(homedir)
    return output

def excel_history_list(history, ws, count_all):

    DATE_COL = "A"
    CATEGORY_COL = "B"
    VALUE_COL = "C"
    USER_COL = "D"
    COLS = [DATE_COL, CATEGORY_COL, VALUE_COL, USER_COL]
    SKIP = 1

    for COLID, COL in enumerate(COLS):
        ws[COL + str(SKIP)] = ["Категория", "Сумма", "Дата", "Внес"][COLID]
        ws[COL + str(SKIP)].font = Font(bold=True)
        column = ws.column_dimensions[COL]
        column.width = [30, 10, 10, 10][COLID]

    for index, entry in enumerate(history.get(count_all)):
        index += (SKIP + 1)
        category = entry[0]
        value = entry[1]
        date = f"{entry[2][2]:0>2}.{entry[2][1]:0>2}.{entry[2][0]:0>4}"
        user = entry[3][1]
        data = [category, value, date, user]
        for COLID, COL in enumerate(COLS):
            ID = COL + str(index)
            if COLID == 1:
                ws[ID].number_format = "#,##0.00"
            elif COLID == 2:
                ws[ID].number_format = "mm.dd.yy"
            ws[ID] = data[COLID]


def excel_expences(history, ws, count_all):

    CATEGORY_COL = "A"
    VALUE_COL = "B"
    COLS = [CATEGORY_COL, VALUE_COL]
    SKIP = 1

    for COLID, COL in enumerate(COLS):
        ws[COL + str(SKIP)] = ["Категория", "Сумма"][COLID]
        ws[COL + str(SKIP)].font = Font(bold=True)
        column = ws.column_dimensions[COL]
        column.width = [30, 10][COLID]

    expences_dict = history.count_detailed(count_all)
    expences_total = history.count_total(count_all)[0]
    income_total = history.count_total(count_all)[1]

    counter = SKIP + 1
    for category, value in sorted(expences_dict.items()):
        ws[CATEGORY_COL + str(counter)] = category
        ws[VALUE_COL + str(counter)] = value
        ws[VALUE_COL + str(counter)].number_format = "#,##0.00"
        counter += 1
    ws[CATEGORY_COL + str(counter)] = "ИТОГО, Доходы"
    ws[CATEGORY_COL + str(counter + 1)] = "ИТОГО, Расходы"
    ws[VALUE_COL + str(counter)] = income_total
    # ws[VALUE_COL + str(counter + 1)] = f"=СУММ({VALUE_COL + str(SKIP)}:{VALUE_COL + str(counter - 1)})"
    ws[VALUE_COL + str(counter + 1)] = expences_total
    ws[VALUE_COL + str(counter)].number_format = "#,#0"
    ws[VALUE_COL + str(counter + 1)].number_format = "#,#0"

    ws[CATEGORY_COL + str(counter)].font = Font(bold=True)
    ws[CATEGORY_COL + str(counter + 1)].font = Font(bold=True)
    ws[VALUE_COL + str(counter)].font = Font(bold=True)
    ws[VALUE_COL + str(counter + 1)].font = Font(bold=True)


def export_to_excel(history, count_all=0):
    month = months(current_month)
    sheetname = f"за {month}"
    filename = f"РАСХОДЫ_{month}.xlsx"
    if count_all != 0:
        sheetname = "TOTAL"
        filename = f"РАСХОДЫ_TOTAL.xlsx"

    wb = xl.Workbook()
    ws_history = wb.active
    ws_history.title = f"История {sheetname}"
    ws_history.sheet_properties.tabColor = "f0d020"
    ws_expences = wb.create_sheet(f"Расходы {sheetname}", 0)
    ws_expences.sheet_properties.tabColor = "cb0e00"
    excel_history_list(history, ws_history, count_all)
    excel_expences(history, ws_expences, count_all)
    wb.save(filename)
    return filename





# CLASSES #

class History:
    def __init__(self):
        try:
            self.history = read_from_file()
        except Exception as e:
            self.history = []
            # log(f"[LOG]: (class History): /__init__: file not found, new history started{e}")

    def add_entry(self, message, date, user):
        result = False
        message = self._check(message)
        if message:
            for _ in [convert_date(date), user]:
                message.append(_)
            self.history.append(str(message))
            try:
                save_to_file(self.history)
            except Exception as e:
                pass
                # log(f"[ERROR]: (class History): /add_entry: file not found{e}")
            finally:
                self.history = read_from_file( )
            result = ["Принято."]
        return result

    def get(self, count_all=0):
        result = self.history
        if not result:
            return [RESPONSE_HISTORY_EMPTY]

        # История за месяц
        start_index = None
        stop_index = None
        if not count_all:
            for i, _ in enumerate(result):
                month = _[2][1]
                if month == current_month:
                    if start_index is None:
                        start_index = i
                        log(f"[DEBUG]: start_index = {start_index}")
                else:
                    if start_index is not None:
                        stop_index = i
                        log(f"[DEBUG]: stop_index = {stop_index}")
                        break
        result = result[start_index:stop_index]
        log(result)

        return result

    def get_by_date(self, count_all=0):
        dates = dict()
        if not self.history:
            return [RESPONSE_HISTORY_EMPTY]
        for _ in self.history:
            date = f"{_[2][2]:0>2}.{_[2][1]:0>2}.{_[2][0]:0>4}"
            value = _[1]
            month = _[2][1]
            if count_all:
                dates.setdefault(date, 0)
                dates[date] += value
            else:
                if month == current_month:
                    dates.setdefault(date, 0)
                    dates[date] += value
        return dates

    def clear(self):
        clear_file()
        self.history = []
        return RESPONSE_HISTORY_CLEARED

    def clear_last(self):
        last_entry = self.history[-1]
        clear_last()
        self.history = read_from_file()
        result = [f"Запись '{last_entry[0]} {last_entry[1]:,.2f} р.' удалена"]
        return result

    def count_total(self, count_all=0):
        summ = 0
        income = 0
        for _ in self.history:
            value = _[1]
            month = _[2][1]
            if count_all:
                if value > 0:
                    income += value
                else:
                    summ += value
            else:
                if month == current_month:
                    # log(f"[LOG]: (class History): /count: value = {value}")
                    if value > 0:
                        income += value
                    else:
                        summ += value
                # else:
                    # log(f"[LOG]: (class History): /count: total line = {value}: month don't match:")
                    # log(f"[LOG]: (class History): /count: current: {current_month}, in line: {month}")
        return summ, income

    def count_detailed(self, count_all=0):
        categories = dict()
        for _ in self.history:
            category = _[0]
            category = re.split(r"\s(?=[A-Z]|[А-Я])", category)[0]
            # log(f"[LOG]: (class History): /count_detailed: category: {category}")
            value = _[1]
            month = _[2][1]
            if count_all:
                if value > 0:
                    continue
                categories.setdefault(category, 0)
                categories[category] += value
            else:
                if month == current_month:
                    # log(f"[LOG]: (class History): /count: category = {category}")
                    if value > 0:
                        continue
                    categories.setdefault(category, 0)
                    categories[category] += value
                # else:
                    # log(f"[LOG]: (class History): /count: line = {category}-{value}: month don't match:")
                    # log(f"[LOG]: (class History): /count: current: {current_month}, in line: {month}")
        return categories

    def count_by_users(self):
        users = dict()
        for _ in self.history:
            user = _[3][1]
            value = _[1]
            month = _[2][1]
            if month == current_month:
                if value < 0:
                    users.setdefault(user, 0)
                    users[user] += value
        return users

    def _check(self, entry):
        result = False
        entry = check_values(entry)
        if entry[0] and entry[1]:
            result = [entry[0], entry[1]]
            # log(f"[LOG]: (class History): /_check: result = [{entry[0].__class__.__name__}, {entry[1].__class__.__name__}]")
        # else:
            # log(f"[ERROR]: (class History): /_check: wrong string format")
        return result

